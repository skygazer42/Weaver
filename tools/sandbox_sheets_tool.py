"""
Sandbox Sheets Tool for E2B Sandbox Spreadsheet Operations.

This module provides spreadsheet generation capabilities in an E2B sandbox:
- Create Excel (.xlsx) and CSV files
- Write data to cells and ranges
- Format cells (bold, colors, borders)
- Create charts (bar, line, pie)
- Add formulas
- Multiple sheet support

Similar to Manus's sb_sheets_tool.py but adapted for Weaver's E2B integration.

Usage:
    from tools.sandbox_sheets_tool import build_sandbox_sheets_tools

    tools = build_sandbox_sheets_tools(thread_id="thread_123")
"""

from __future__ import annotations

import asyncio
import base64
import json
import logging
import time
from typing import Any, Dict, List, Optional, Union

from langchain_core.tools import BaseTool
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


def _get_sandbox_session(thread_id: str):
    """Get sandbox session for a thread."""
    try:
        from tools.sandbox_browser_session import sandbox_browser_sessions
        return sandbox_browser_sessions.get(thread_id)
    except ImportError:
        return None


def _get_event_emitter(thread_id: str):
    """Get event emitter for a thread."""
    try:
        from agent.events import get_emitter_sync
        return get_emitter_sync(thread_id)
    except ImportError:
        return None


class _SandboxSheetsBaseTool(BaseTool):
    """Base class for sandbox sheets tools."""

    thread_id: str = "default"
    emit_events: bool = True
    workspace_path: str = "/workspace"

    def _get_sandbox(self):
        """Get the E2B sandbox instance."""
        session = _get_sandbox_session(self.thread_id)
        if session and hasattr(session, "_handles") and session._handles:
            return session._handles.sandbox
        return None

    def _emit_event(self, event_type: str, data: Dict[str, Any]) -> None:
        """Emit an event."""
        if not self.emit_events:
            return
        emitter = _get_event_emitter(self.thread_id)
        if emitter:
            try:
                loop = asyncio.new_event_loop()
                try:
                    loop.run_until_complete(emitter.emit(event_type, data))
                finally:
                    loop.close()
            except Exception as e:
                logger.warning(f"[sandbox_sheets] Failed to emit event: {e}")

    def _emit_tool_start(self, action: str, args: Dict[str, Any]) -> float:
        """Emit tool start event."""
        start_time = time.time()
        self._emit_event("tool_start", {
            "tool": self.name,
            "action": action,
            "args": args,
            "thread_id": self.thread_id,
        })
        return start_time

    def _emit_tool_result(
        self,
        action: str,
        result: Dict[str, Any],
        start_time: float,
        success: bool = True,
    ) -> None:
        """Emit tool result event."""
        duration_ms = (time.time() - start_time) * 1000
        self._emit_event("tool_result", {
            "tool": self.name,
            "action": action,
            "success": success,
            "duration_ms": round(duration_ms, 2),
        })

    def _ensure_openpyxl(self, sandbox) -> bool:
        """Ensure openpyxl is installed in sandbox."""
        try:
            result = sandbox.commands.run("pip show openpyxl", timeout=30)
            if result.exit_code != 0:
                logger.info("[sandbox_sheets] Installing openpyxl...")
                install_result = sandbox.commands.run(
                    "pip install openpyxl xlsxwriter pandas",
                    timeout=120
                )
                return install_result.exit_code == 0
            return True
        except Exception as e:
            logger.warning(f"[sandbox_sheets] Failed to check/install openpyxl: {e}")
            return False


class CreateSpreadsheetInput(BaseModel):
    """Input for create_spreadsheet."""
    file_path: str = Field(
        description="Path for the spreadsheet file (e.g., 'reports/data.xlsx' or 'output.csv')"
    )
    sheet_name: str = Field(
        default="Sheet1",
        description="Name of the initial sheet"
    )
    headers: Optional[List[str]] = Field(
        default=None,
        description="Optional list of column headers"
    )


class SandboxCreateSpreadsheetTool(_SandboxSheetsBaseTool):
    """Create a new spreadsheet file."""

    name: str = "sandbox_create_spreadsheet"
    description: str = (
        "Create a new Excel (.xlsx) or CSV spreadsheet file. "
        "Optionally set the initial sheet name and column headers. "
        "Path must be relative to /workspace."
    )
    args_schema: type[BaseModel] = CreateSpreadsheetInput

    def _run(
        self,
        file_path: str,
        sheet_name: str = "Sheet1",
        headers: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("create_spreadsheet", {"file_path": file_path})

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized. Start sandbox browser first.")

            # Ensure openpyxl is installed
            if not self._ensure_openpyxl(sandbox):
                return {"success": False, "error": "Failed to install openpyxl"}

            # Determine file type
            is_csv = file_path.lower().endswith(".csv")
            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"

            # Create parent directories
            parent_dir = "/".join(full_path.split("/")[:-1])
            if parent_dir:
                sandbox.commands.run(f"mkdir -p {parent_dir}")

            if is_csv:
                # Create CSV file
                if headers:
                    header_line = ",".join(f'"{h}"' for h in headers)
                    sandbox.filesystem.write(full_path, header_line + "\n")
                else:
                    sandbox.filesystem.write(full_path, "")
            else:
                # Create Excel file with Python
                headers_json = json.dumps(headers) if headers else "None"
                python_code = f'''
import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "{sheet_name}"

headers = {headers_json}
if headers:
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        # Make header bold
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

wb.save("{full_path}")
print("SUCCESS")
'''
                result = sandbox.commands.run(f'python3 -c "{python_code}"', timeout=30)
                if "SUCCESS" not in result.stdout:
                    raise RuntimeError(f"Failed to create spreadsheet: {result.stderr}")

            result = {
                "success": True,
                "message": f"Spreadsheet created: {file_path}",
                "path": file_path,
                "type": "csv" if is_csv else "xlsx",
                "sheet_name": sheet_name if not is_csv else None,
            }

            self._emit_tool_result("create_spreadsheet", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("create_spreadsheet", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


class WriteDataInput(BaseModel):
    """Input for write_data."""
    file_path: str = Field(description="Path to the spreadsheet file")
    data: List[List[Any]] = Field(
        description="2D array of data to write (rows x columns)"
    )
    start_row: int = Field(default=1, description="Starting row (1-based)")
    start_col: int = Field(default=1, description="Starting column (1-based)")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name (for Excel)")


class SandboxWriteDataTool(_SandboxSheetsBaseTool):
    """Write data to spreadsheet cells."""

    name: str = "sandbox_write_data"
    description: str = (
        "Write data to a spreadsheet. Provide a 2D array of values. "
        "Specify starting row/column (1-based indexing). "
        "For Excel files, optionally specify the sheet name."
    )
    args_schema: type[BaseModel] = WriteDataInput

    def _run(
        self,
        file_path: str,
        data: List[List[Any]],
        start_row: int = 1,
        start_col: int = 1,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("write_data", {
            "file_path": file_path,
            "rows": len(data),
            "cols": len(data[0]) if data else 0,
        })

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized.")

            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"
            is_csv = file_path.lower().endswith(".csv")

            if is_csv:
                # Append to CSV
                lines = []
                for row in data:
                    line = ",".join(
                        f'"{str(cell)}"' if "," in str(cell) else str(cell)
                        for cell in row
                    )
                    lines.append(line)

                try:
                    existing = sandbox.filesystem.read(full_path)
                except Exception:
                    existing = ""

                new_content = existing.rstrip("\n") + "\n" + "\n".join(lines) + "\n"
                sandbox.filesystem.write(full_path, new_content)
            else:
                # Write to Excel
                data_json = json.dumps(data)
                sheet_arg = f'"{sheet_name}"' if sheet_name else "None"
                python_code = f'''
import openpyxl
import json

wb = openpyxl.load_workbook("{full_path}")
sheet_name = {sheet_arg}
if sheet_name:
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
else:
    ws = wb.active

data = json.loads('{data_json}')
start_row = {start_row}
start_col = {start_col}

for r_idx, row in enumerate(data):
    for c_idx, value in enumerate(row):
        ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

wb.save("{full_path}")
print("SUCCESS")
'''
                result = sandbox.commands.run(f"python3 -c '{python_code}'", timeout=30)
                if "SUCCESS" not in result.stdout:
                    raise RuntimeError(f"Failed to write data: {result.stderr}")

            result = {
                "success": True,
                "message": f"Data written: {len(data)} rows",
                "path": file_path,
                "rows_written": len(data),
                "start_row": start_row,
                "start_col": start_col,
            }

            self._emit_tool_result("write_data", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("write_data", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


class FormatCellsInput(BaseModel):
    """Input for format_cells."""
    file_path: str = Field(description="Path to the Excel file")
    start_row: int = Field(description="Starting row (1-based)")
    end_row: int = Field(description="Ending row (1-based)")
    start_col: int = Field(description="Starting column (1-based)")
    end_col: int = Field(description="Ending column (1-based)")
    bold: bool = Field(default=False, description="Make text bold")
    italic: bool = Field(default=False, description="Make text italic")
    font_size: Optional[int] = Field(default=None, description="Font size")
    font_color: Optional[str] = Field(default=None, description="Font color (hex, e.g., 'FF0000')")
    bg_color: Optional[str] = Field(default=None, description="Background color (hex)")
    border: bool = Field(default=False, description="Add cell borders")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name")


class SandboxFormatCellsTool(_SandboxSheetsBaseTool):
    """Format cells in an Excel spreadsheet."""

    name: str = "sandbox_format_cells"
    description: str = (
        "Apply formatting to a range of cells in an Excel file. "
        "Supports bold, italic, colors, borders, and font size. "
        "Only works with .xlsx files."
    )
    args_schema: type[BaseModel] = FormatCellsInput

    def _run(
        self,
        file_path: str,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        bold: bool = False,
        italic: bool = False,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        border: bool = False,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("format_cells", {
            "file_path": file_path,
            "range": f"({start_row},{start_col}) to ({end_row},{end_col})",
        })

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized.")

            if file_path.lower().endswith(".csv"):
                return {"success": False, "error": "Formatting only supported for Excel files"}

            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"
            sheet_arg = f'"{sheet_name}"' if sheet_name else "None"

            python_code = f'''
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

wb = openpyxl.load_workbook("{full_path}")
sheet_name = {sheet_arg}
ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

font_kwargs = {{}}
if {bold}:
    font_kwargs["bold"] = True
if {italic}:
    font_kwargs["italic"] = True
if {font_size}:
    font_kwargs["size"] = {font_size}
if "{font_color}" and "{font_color}" != "None":
    font_kwargs["color"] = "{font_color}"

fill = None
if "{bg_color}" and "{bg_color}" != "None":
    fill = PatternFill(start_color="{bg_color}", end_color="{bg_color}", fill_type="solid")

border_style = None
if {border}:
    side = Side(style="thin")
    border_style = Border(left=side, right=side, top=side, bottom=side)

for row in range({start_row}, {end_row} + 1):
    for col in range({start_col}, {end_col} + 1):
        cell = ws.cell(row=row, column=col)
        if font_kwargs:
            cell.font = Font(**font_kwargs)
        if fill:
            cell.fill = fill
        if border_style:
            cell.border = border_style

wb.save("{full_path}")
print("SUCCESS")
'''
            result = sandbox.commands.run(f"python3 -c '{python_code}'", timeout=30)
            if "SUCCESS" not in result.stdout:
                raise RuntimeError(f"Failed to format cells: {result.stderr}")

            result = {
                "success": True,
                "message": "Cells formatted successfully",
                "path": file_path,
                "range": f"({start_row},{start_col}) to ({end_row},{end_col})",
            }

            self._emit_tool_result("format_cells", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("format_cells", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


class CreateChartInput(BaseModel):
    """Input for create_chart."""
    file_path: str = Field(description="Path to the Excel file")
    chart_type: str = Field(
        description="Chart type: 'bar', 'line', 'pie', 'scatter', 'area'"
    )
    data_range: str = Field(
        description="Data range in A1 notation (e.g., 'A1:D10')"
    )
    title: str = Field(default="", description="Chart title")
    x_axis_title: str = Field(default="", description="X-axis title")
    y_axis_title: str = Field(default="", description="Y-axis title")
    position: str = Field(default="E1", description="Chart position (cell reference)")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name")


class SandboxCreateChartTool(_SandboxSheetsBaseTool):
    """Create a chart in an Excel spreadsheet."""

    name: str = "sandbox_create_chart"
    description: str = (
        "Create a chart in an Excel file. "
        "Supports bar, line, pie, scatter, and area charts. "
        "Specify data range in A1 notation (e.g., 'A1:D10')."
    )
    args_schema: type[BaseModel] = CreateChartInput

    def _run(
        self,
        file_path: str,
        chart_type: str,
        data_range: str,
        title: str = "",
        x_axis_title: str = "",
        y_axis_title: str = "",
        position: str = "E1",
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("create_chart", {
            "file_path": file_path,
            "chart_type": chart_type,
            "data_range": data_range,
        })

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized.")

            if file_path.lower().endswith(".csv"):
                return {"success": False, "error": "Charts only supported for Excel files"}

            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"
            sheet_arg = f'"{sheet_name}"' if sheet_name else "None"

            chart_map = {
                "bar": "BarChart",
                "line": "LineChart",
                "pie": "PieChart",
                "scatter": "ScatterChart",
                "area": "AreaChart",
            }
            chart_class = chart_map.get(chart_type.lower(), "BarChart")

            python_code = f'''
import openpyxl
from openpyxl.chart import {chart_class}, Reference

wb = openpyxl.load_workbook("{full_path}")
sheet_name = {sheet_arg}
ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

chart = {chart_class}()
chart.title = "{title}"
chart.x_axis.title = "{x_axis_title}"
chart.y_axis.title = "{y_axis_title}"

# Parse data range
data_range = "{data_range}"
# Assuming format like A1:D10
import re
match = re.match(r"([A-Z]+)(\\d+):([A-Z]+)(\\d+)", data_range)
if match:
    start_col = ord(match.group(1)) - ord('A') + 1
    start_row = int(match.group(2))
    end_col = ord(match.group(3)) - ord('A') + 1
    end_row = int(match.group(4))

    data = Reference(ws, min_col=start_col+1, min_row=start_row, max_col=end_col, max_row=end_row)
    categories = Reference(ws, min_col=start_col, min_row=start_row+1, max_row=end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

ws.add_chart(chart, "{position}")
wb.save("{full_path}")
print("SUCCESS")
'''
            result = sandbox.commands.run(f"python3 -c '{python_code}'", timeout=30)
            if "SUCCESS" not in result.stdout:
                raise RuntimeError(f"Failed to create chart: {result.stderr}")

            result = {
                "success": True,
                "message": f"Chart created: {chart_type}",
                "path": file_path,
                "chart_type": chart_type,
                "data_range": data_range,
                "position": position,
            }

            self._emit_tool_result("create_chart", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("create_chart", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


class AddSheetInput(BaseModel):
    """Input for add_sheet."""
    file_path: str = Field(description="Path to the Excel file")
    sheet_name: str = Field(description="Name for the new sheet")
    position: Optional[int] = Field(default=None, description="Position index (0-based)")


class SandboxAddSheetTool(_SandboxSheetsBaseTool):
    """Add a new sheet to an Excel workbook."""

    name: str = "sandbox_add_sheet"
    description: str = (
        "Add a new sheet to an existing Excel workbook. "
        "Optionally specify the position (0 = first)."
    )
    args_schema: type[BaseModel] = AddSheetInput

    def _run(
        self,
        file_path: str,
        sheet_name: str,
        position: Optional[int] = None,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("add_sheet", {
            "file_path": file_path,
            "sheet_name": sheet_name,
        })

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized.")

            if file_path.lower().endswith(".csv"):
                return {"success": False, "error": "Multiple sheets only supported for Excel files"}

            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"
            position_arg = str(position) if position is not None else "None"

            python_code = f'''
import openpyxl

wb = openpyxl.load_workbook("{full_path}")
position = {position_arg}

if position is not None:
    wb.create_sheet("{sheet_name}", position)
else:
    wb.create_sheet("{sheet_name}")

wb.save("{full_path}")
print("SUCCESS")
print(f"Sheets: {{wb.sheetnames}}")
'''
            result = sandbox.commands.run(f"python3 -c '{python_code}'", timeout=30)
            if "SUCCESS" not in result.stdout:
                raise RuntimeError(f"Failed to add sheet: {result.stderr}")

            # Extract sheet names from output
            sheets = []
            for line in result.stdout.split("\n"):
                if line.startswith("Sheets:"):
                    try:
                        sheets = eval(line.replace("Sheets:", "").strip())
                    except Exception:
                        pass

            result = {
                "success": True,
                "message": f"Sheet '{sheet_name}' added",
                "path": file_path,
                "sheet_name": sheet_name,
                "all_sheets": sheets,
            }

            self._emit_tool_result("add_sheet", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("add_sheet", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


class ReadSpreadsheetInput(BaseModel):
    """Input for read_spreadsheet."""
    file_path: str = Field(description="Path to the spreadsheet file")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name (for Excel)")
    max_rows: int = Field(default=100, description="Maximum rows to read")
    start_row: int = Field(default=1, description="Starting row (1-based)")


class SandboxReadSpreadsheetTool(_SandboxSheetsBaseTool):
    """Read data from a spreadsheet."""

    name: str = "sandbox_read_spreadsheet"
    description: str = (
        "Read data from a spreadsheet file (Excel or CSV). "
        "Returns data as a 2D array. Use max_rows to limit output."
    )
    args_schema: type[BaseModel] = ReadSpreadsheetInput

    def _run(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        max_rows: int = 100,
        start_row: int = 1,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("read_spreadsheet", {"file_path": file_path})

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized.")

            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"
            is_csv = file_path.lower().endswith(".csv")

            if is_csv:
                # Read CSV
                content = sandbox.filesystem.read(full_path)
                lines = content.strip().split("\n")
                data = []
                for line in lines[start_row-1:start_row-1+max_rows]:
                    # Simple CSV parsing
                    row = []
                    in_quote = False
                    current = ""
                    for char in line:
                        if char == '"':
                            in_quote = not in_quote
                        elif char == "," and not in_quote:
                            row.append(current.strip('"'))
                            current = ""
                        else:
                            current += char
                    row.append(current.strip('"'))
                    data.append(row)
            else:
                # Read Excel
                sheet_arg = f'"{sheet_name}"' if sheet_name else "None"
                python_code = f'''
import openpyxl
import json

wb = openpyxl.load_workbook("{full_path}", data_only=True)
sheet_name = {sheet_arg}
ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

data = []
for row_idx, row in enumerate(ws.iter_rows(min_row={start_row}, max_row={start_row + max_rows - 1})):
    row_data = []
    for cell in row:
        value = cell.value
        if value is None:
            value = ""
        row_data.append(str(value) if value != "" else "")
    data.append(row_data)

print(json.dumps({{"data": data, "sheets": wb.sheetnames}}))
'''
                result = sandbox.commands.run(f"python3 -c '{python_code}'", timeout=30)
                try:
                    output = json.loads(result.stdout.strip())
                    data = output["data"]
                except Exception:
                    raise RuntimeError(f"Failed to parse spreadsheet: {result.stderr}")

            result = {
                "success": True,
                "path": file_path,
                "data": data,
                "rows": len(data),
                "columns": len(data[0]) if data else 0,
            }

            self._emit_tool_result("read_spreadsheet", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("read_spreadsheet", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


class AddFormulaInput(BaseModel):
    """Input for add_formula."""
    file_path: str = Field(description="Path to the Excel file")
    cell: str = Field(description="Cell reference (e.g., 'E1')")
    formula: str = Field(description="Excel formula (e.g., '=SUM(A1:A10)')")
    sheet_name: Optional[str] = Field(default=None, description="Sheet name")


class SandboxAddFormulaTool(_SandboxSheetsBaseTool):
    """Add a formula to an Excel cell."""

    name: str = "sandbox_add_formula"
    description: str = (
        "Add an Excel formula to a cell. "
        "Formula should start with '=' (e.g., '=SUM(A1:A10)'). "
        "Only works with .xlsx files."
    )
    args_schema: type[BaseModel] = AddFormulaInput

    def _run(
        self,
        file_path: str,
        cell: str,
        formula: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        start_time = self._emit_tool_start("add_formula", {
            "file_path": file_path,
            "cell": cell,
            "formula": formula,
        })

        try:
            sandbox = self._get_sandbox()
            if not sandbox:
                raise RuntimeError("Sandbox not initialized.")

            if file_path.lower().endswith(".csv"):
                return {"success": False, "error": "Formulas only supported for Excel files"}

            full_path = f"{self.workspace_path}/{file_path.lstrip('/')}"
            sheet_arg = f'"{sheet_name}"' if sheet_name else "None"

            # Escape formula for Python string
            formula_escaped = formula.replace("'", "\\'")

            python_code = f'''
import openpyxl

wb = openpyxl.load_workbook("{full_path}")
sheet_name = {sheet_arg}
ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

ws["{cell}"] = '{formula_escaped}'

wb.save("{full_path}")
print("SUCCESS")
'''
            result = sandbox.commands.run(f"python3 -c \"{python_code}\"", timeout=30)
            if "SUCCESS" not in result.stdout:
                raise RuntimeError(f"Failed to add formula: {result.stderr}")

            result = {
                "success": True,
                "message": f"Formula added to cell {cell}",
                "path": file_path,
                "cell": cell,
                "formula": formula,
            }

            self._emit_tool_result("add_formula", result, start_time, True)
            return result

        except Exception as e:
            self._emit_tool_result("add_formula", {"error": str(e)}, start_time, False)
            return {"success": False, "error": str(e)}


def build_sandbox_sheets_tools(
    thread_id: str,
    emit_events: bool = True,
) -> List[BaseTool]:
    """
    Build sandbox spreadsheet tools for a thread.

    Args:
        thread_id: Thread/conversation ID
        emit_events: Whether to emit events

    Returns:
        List of spreadsheet tools
    """
    return [
        SandboxCreateSpreadsheetTool(thread_id=thread_id, emit_events=emit_events),
        SandboxWriteDataTool(thread_id=thread_id, emit_events=emit_events),
        SandboxReadSpreadsheetTool(thread_id=thread_id, emit_events=emit_events),
        SandboxFormatCellsTool(thread_id=thread_id, emit_events=emit_events),
        SandboxAddFormulaTool(thread_id=thread_id, emit_events=emit_events),
        SandboxCreateChartTool(thread_id=thread_id, emit_events=emit_events),
        SandboxAddSheetTool(thread_id=thread_id, emit_events=emit_events),
    ]
